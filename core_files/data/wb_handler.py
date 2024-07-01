import sys
import re
from openpyxl import load_workbook
from data.data_handler import Data_Handler

def generate_spreadsheet_name (model, replacements_list):
    if model == 'Component':
        name_key = '{component_name}'
        type_key = '{component_type}'

        for key, value in replacements_list:
            if key == name_key:
                name_value = value.replace(" ","_")
            if key == type_key:
                type_value = value
    else:
        name_key = '{feature_name}'
        type_key = '{feature_type}'

        for key, value in replacements_list:
            if key == name_key:
                name_value = value.replace(" ","_")
            if key == type_key:
                type_value = value

    return name_value, type_value


def generate_output_spreadsheet (tool, model, replacements_list):
    name_value, type_value = generate_spreadsheet_name (model, replacements_list)
    output_spreadsheet = Data_Handler.copy_config_workbook(tool, model, name_value, type_value)
    wb = load_workbook(output_spreadsheet)
    ws = wb.active
    ws.delete_rows

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str):
                for key, value in replacements_list:
                    cell.value = cell.value.replace(key, value)
                cell.value = re.sub(r"\s*\{other_pre_condition_\d+\}\s*", "\n", cell.value).strip()
                lines = cell.value.split("\n")
                updated_lines = [line.strip() for line in lines if line.strip()]
                cell.value = "\n".join(updated_lines)            
    
    wb.save(output_spreadsheet)




class WB_Handler:

    def generate_output_spreadsheets (selected_tools, selected_models, replacement_file):
        replacements_json = Data_Handler.map_features_json_to_features_dict(replacement_file)

        for tool in selected_tools: # ADO, Jira...
            for model in selected_models: # Feature, Components... 
                for feature in replacements_json: #Explore screen, Rewards screen...    
                    replacements_list = []
                    for key, value in feature.items(): # i.e.
                        if key not in ("components"):
                            replacements_list.append([key, value])
                    if model == 'Component':
                        if 'components' in feature and feature['components']: 
                            for component in feature['components']:  # Access components within the feature
                                original_length = len(replacements_list)
                                for key, value in component.items():
                                    replacements_list.append([key,value])
                                generate_output_spreadsheet(tool, model, replacements_list)
                                replacements_list = replacements_list[:original_length]
                        else:
                            pass
                    else:
                        generate_output_spreadsheet (tool, model, replacements_list)